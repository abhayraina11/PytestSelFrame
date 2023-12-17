import openpyxl

book = openpyxl.load_workbook("C:\\Users\\HP\\Documents\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
for i in range(1,sheet.max_row+1):
    print(sheet.cell(row=i,column=1).value)